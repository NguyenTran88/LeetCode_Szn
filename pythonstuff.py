import openpyxl as xl  # give the package a shorter alias
from openpyxl.chart import BarChart, Reference


def fix_workbook(self, filename):
    # the load_wb method returns a workbook that has the name transction.xlsx
    wb = xl.load_workbook('filename.xlsx')
    sheet = wb['Sheet1']

    # 2 bc 1st row k co gi, n+1 since range() is exclusive
    for row in range(2, sheet.max_row + 1):
        mycell = sheet.cell(row, 3)
        correct_cell = sheet.cell(row, 5)
        correct_cell.value = mycell.value * 0.9

    values = Reference(sheet, min_row=2,
                       max_row=sheet.max_row,
                       min_col=5,
                       max_col=5)
    # reference gives back all the cells in the 4*4 range we choose

    chart = BarChart()  # create chart
    chart.add_data(values)  # fill value
    # add chart to sheet with e2 as topleft corner of chart
    sheet.add_chart(chart, 'e2')

    wb.save('filename.xlxs')  # overwrite if exists, new file if not
